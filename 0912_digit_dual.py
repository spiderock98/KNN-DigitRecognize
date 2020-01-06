import cv2
import serial
import numpy as np
from time import sleep
from threading import Thread
import portusb
import cameraindex
import subprocess
from openpyxl import load_workbook, Workbook
import datetime
import os
import shutil
import glob
from zoom import Zoom
from imutils.video import FPS

######################## checking MAC address #########################

# mac=open('/sys/class/net/%s/address' %'eth0').read().strip()
# while( mac!= 'b8:27:eb:df:bd:66'):
#     continue

############################ Training  ############################

samples = np.loadtxt('generalsamples4.data',np.float32)
responses = np.loadtxt('generalresponses4.data',np.float32)
responses = responses.reshape((responses.size,1))
model = cv2.ml.KNearest_create()
model.train(samples, cv2.ml.ROW_SAMPLE, responses)

############################# Init Vars #########################

dctChar3 = dctChar4 = dict() #[vi_tri_x:value]
preRet3 = preRet4 = iFrame3 = iFrame4 = iFail3 = iFail4 = finalReceive = screenRet4 = iInit = lowfinalRec4 = upfinalRec4 = finalRec3 = 0
final3 = final4 = False
flagPass = flagFail = flagSpeed = flagInit = False
seriFrame = 20
# sentFlag = False
# cam4Flag = True

############################# Function  #########################

def newWorkbook():
    global wb,ws
    wb = Workbook()
    ws = wb.active
    # title bar
    timeCol = ws.cell(row=1,column=2,value='Time')
    timeCol.style = 'Title'
    sttCol = ws.cell(row=1,column=3,value='Status')
    sttCol.style = 'Title'
    errRow = ws.cell(row=4,column=6,value='Error')
    errRow.style = 'Calculation'
    totalRow = ws.cell(row=5,column=6,value='Total')
    totalRow.style = 'Calculation'
def insertRow(status):
    global wb,ws
    ws.append([None,datetime.datetime.now().strftime('%d-%m-%Y %H:%M:%S'),status])
    if (status == 'Good'):
        ws['C{}'.format(ws.max_row)].style = 'Good'
    else:
        ws['C{}'.format(ws.max_row)].style = 'Bad'
    wb.save('logs.xlsx')
def filter():
    ws.auto_filter.ref = 'A1:C{}'.format(ws.max_row)
    wb.save('logs.xlsx')

############################# Init I/O #########################

if (os.path.isfile('./logs.xlsx')):
    wb = load_workbook(filename='logs.xlsx')
    ws = wb.active
else:
    newWorkbook()

#port = portusb.port()
port = portusb.portNew()
print('[INFO] Openning ' + port)
ser = serial.Serial(port,9600)

vs3 = Zoom(400,cameraindex.index('Camera'),169,266,189,377).start()
vs4 = Zoom(400,cameraindex.index('CAMERA'),263,372,179,438).start()

sleep(2.0)

fps = FPS().start()

############################# Loop #########################

while True:
    try:
        serInWait = ser.in_waiting
    except OSError:
        vs3.stop()
        vs4.stop()
        ser.close()
        cv2.destroyAllWindows()
        os.system('python3 ./0908_digit_dual.py')
    if (serInWait):
        recieveOrd = [] # reset
        for i in range(13):
            capture = ord(ser.read(1))
            if (capture == 110):
                print('[INFO] Resetting arduino ...')
                screenRet4 = iInit = 0
                final3 = final4 = flagInit = flagFail = flagPass = False
                flagSpeed = True
                ### chi test thu o day ####
                #insertRow('Bad')
            if (capture == 111):
                print('[INFO] Detecting valid USB')
                strUSB = str(subprocess.check_output('lsusb',shell=True))
                #strUSB = glob.glob('/media/pi/[A-Za-z]*')[0]
                if strUSB.rfind('0951:1666') != -1:
                    print('[INFO] USB access grandted')
                    ser.write('USB'.encode())
                    #auto filter
                    ws.auto_filter.ref = 'A1:C{}'.format(ws.max_row)
                    wb.save('logs.xlsx')
                    # cut it out
                    if(os.path.isfile('./logs.xlsx')):
                        shutil.copy('./logs.xlsx', glob.glob('/media/pi/[A-Za-z]*')[0])
                        sleep(5) # waiting for copy
                        os.remove('./logs.xlsx')
                    # then create new wb
                    newWorkbook()
                    ser.write('usb'.encode())
                else:
                    print('[INFO] USB access denied')
            recieveOrd.append(capture)
        # check error nếu bit 3 != 30, 31 và != tổng 3 bit còn lại -> lỗi
        if(recieveOrd[0]+recieveOrd[1]+recieveOrd[2]+recieveOrd[4]+recieveOrd[5]+recieveOrd[6]+recieveOrd[7]+recieveOrd[8]+recieveOrd[9]+recieveOrd[10]+recieveOrd[11])!=recieveOrd[12]:
            ser.write("RESEND\r".encode())
            sleep(0.5)
        finalRec3 = float((recieveOrd[0])*100 + (recieveOrd[1])*10 + (recieveOrd[2]))/100
        lowfinalRec4 = recieveOrd[4]*1000 + recieveOrd[5]*100 + recieveOrd[6]*10 + recieveOrd[7]
        upfinalRec4 = recieveOrd[8]*1000 + recieveOrd[9]*100 + recieveOrd[10]*10 + recieveOrd[11]
        print('[INFO] Thresh: {}'.format(finalRec3))
        print('[INFO] Lower Speed: {} \t Upper Speed: {}'.format(lowfinalRec4, upfinalRec4))

    vs4.cont()
    vs3.pause()
    dst4 = vs4.dst
    roi4 = vs4.zoom
    roiBar4 = vs4.roiBar
    cv2.imshow('Zoom{}'.format(vs4.src), dst4)
    #cv2.imshow('Panel{}'.format(vs4.src), roi4)
    contours4,_= cv2.findContours(dst4,cv2.RETR_EXTERNAL,cv2.CHAIN_APPROX_SIMPLE)
    dctChar4 = dict() #reset
    for cnt in contours4:
        [x4,y4,w4,h4] = cv2.boundingRect(cnt)
        if (w4 > roiBar4 and h4 > roiBar4):
            _,results,_,_ = model.findNearest(np.float32(cv2.resize(dst4[y4:y4+h4,x4:x4+w4],(10,10)).reshape((1,100))), k = 5)
            dctChar4[x4] = int(results[0][0])
            cv2.rectangle(roi4,(x4,y4),(x4+w4,y4+h4),(0,0,255),2)

    if(len(dctChar4) == 4):
        multi = []
        for xCoor in sorted(dctChar4):
            multi.append(dctChar4[xCoor])
        screenRet4 = multi[0]*1000 + multi[1]*100 + multi[2]*10 + multi[3]
        #if (screenRet4 != preRet4 ):
            #preRet4 = screenRet4
            #print('\t{}'.format(screenRet4))
            
        if ((screenRet4 != preRet4) and (not flagInit)):
            preRet4 = screenRet4
            iInit += 1
            if (iInit > 4):
                iInit = 0
                flagInit = True
                flagSpeed = False
            print('\t{}'.format(screenRet4))
        if (screenRet4 >= lowfinalRec4 and screenRet4 <= upfinalRec4):
            iFrame4 += 1
            #iFail4 = 0
            if (iFrame4 > seriFrame):
                # cam4Flag = False
                # sleep(3)
                # print('[INFO] speed ok')
                # ser.write('SPEED\r'.encode())
                # sleep(0.5)
                final4 = True # red speed ok
                iFrame4 = 0
        else:
            iFrame4=0
        if ((screenRet4 > 1500 and screenRet4 < lowfinalRec4) or screenRet4 > upfinalRec4):
            iFail4 += 1
            #iFrame4 = 0
            if (iFail4 > seriFrame):
                final4 = False
                iFail4 = 0
        else:
            iFail4=0

    vs3.cont()
    vs4.pause()
    dst3 = vs3.dst
    roi3 = vs3.zoom
    roiBar3 = vs3.roiBar
    cv2.imshow('Zoom{}'.format(vs3.src),dst3)
    #cv2.imshow('Panel{}'.format(vs3.src),roi3)
    contours3,_= cv2.findContours(dst3,cv2.RETR_EXTERNAL,cv2.CHAIN_APPROX_SIMPLE)
    dctChar3 = dict() #reset
    for cnt in contours3:
        [x3,y3,w3,h3] = cv2.boundingRect(cnt)
        if (w3 > roiBar3 and h3 > roiBar3):
            _,results,_,_ = model.findNearest(np.float32(cv2.resize(dst3[y3:y3+h3,x3:x3+w3],(10,10)).reshape((1,100))), k = 5)
            dctChar3[x3] = int(results[0][0])
            cv2.rectangle(roi3,(x3,y3),(x3+w3,y3+h3),(0,0,255),2)

    if (len(dctChar3) == 3):
        multi = []
        for xCoor in sorted(dctChar3):
            multi.append(dctChar3[xCoor])
        screenRet3 = float(multi[0]*100 + multi[1]*10 + multi[2])/100
        if (screenRet3 != preRet3):
            preRet3 = screenRet3
            print(screenRet3)
        if (screenRet3 <= finalRec3):
            iFail3 = 0
            iFrame3 += 1
            if (iFrame3 > seriFrame):
                # print('[INFO] PASS')
                # insertRow('Good')
                # ser.write("PASS\r".encode()) # green light
                # sleep(0.5)
                final3 = True
                iFrame3 = 0
        else:
            iFrame3 = 0
            iFail3 += 1
            if (iFail3 > seriFrame):
                final3 = False
                iFail3 = 0

    if (final3 and final4 and not flagPass):
        flagPass = True
        flagSpeed = False
        print('[INFO] Pass')
        ser.write('PASS\r'.encode())
    elif (screenRet4 <= 1500 and not flagFail):
        flagFail = flagSpeed = True
        flagPass = flagInit = False
        print('[INFO] WHITE')
        ser.write('WHITE\r'.encode())
    elif (not final3 and final4 and not flagSpeed):
        flagSpeed = True
        flagPass = False
        print('[INFO] Failed 1')
        ser.write('SPEED\r'.encode())
    elif (final3 and not final4 and not flagSpeed):
        flagSpeed = True
        flagPass = False
        print('[INFO] Failed 2')
        ser.write('SPEED\r'.encode())
    elif (not final3 and not final4 and not flagSpeed):
        flagSpeed = True
        flagPass = False
        print('[INFO] Failed 3')
        ser.write('SPEED\r'.encode())
    
    if cv2.waitKey(1) == 27:
        break
    fps.update()

fps.stop()
print('[INFO] FPS: {:.2f}'.format(fps.fps()))
vs3.stop()
vs4.stop()
ser.close()
cv2.destroyAllWindows()
